import openpyxl

wb = openpyxl.load_workbook('./datas/example.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')
#从表中取得单元格
values = sheet['A1'].value
print(values)
c = sheet['B1']
print(c.value)
stra = 'Row ' + str(c.row) + ', Column ' + c.column + ' is ' + c.value
print(stra)

strb = 'Cell ' + c.coordinate + ' is ' + c.value
print(strb)

print(sheet['C1'].value)

print(sheet.cell(row=1, column=2))

print(sheet.cell(row=1, column=2).value)
print('=============================')
for i in range(1,8,2):
    print(i,sheet.cell(row=i,column=2).value)


#从表中取得行和列
print('=============================')
print(tuple(sheet['A1':'C3']))
print('=============================')

for row in sheet['A1':'C3']:
    for cell in row:
        print(cell.coordinate,cell.value)
    print('--- END OF ROW ---')


